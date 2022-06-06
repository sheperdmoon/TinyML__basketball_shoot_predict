import numpy
import openpyxl
import pandas

data = pandas.read_excel('ljr&xmh_cosine.xlsx',sheet_name=0)  # 打开xls⽂件
sun_list = numpy.zeros((5,5))
for i in range(0, 25):  # 循环逐⾏打印
    for j in range(0,25):
        sun_list[int((i) / 5)][int((j) / 5)] += abs(data.values[i,j])

for i in range(5):
    for j in range(5):
        print(round(sun_list[i][j],2),end=" ")
    print()

f = openpyxl.Workbook()
sheet1 = f.create_sheet()  # 创建sheet
# 将数据写⼊第 i ⾏，第 j 列
i = 1
for data in sun_list:
    for j in range(len(data)):
        sheet1.cell(i, j+1, data[j])
    i = i + 1
f.save("now.xlsx")  # 保存⽂件

data = pandas.read_excel('data_x_dis.xlsx',sheet_name=0)  # 打开xls⽂件
sun_list = numpy.zeros((5,5))
for i in range(0, 25):  # 循环逐⾏打印
    for j in range(0,25):
        sun_list[int((i) / 5)][int((j) / 5)] += abs(data.values[i,j]/10000)
f = openpyxl.Workbook()
sheet1 = f.create_sheet()  # 创建sheet
# 将数据写⼊第 i ⾏，第 j 列
i = 1
for data in sun_list:
    for j in range(len(data)):
        sheet1.cell(i, j+1, data[j])
    i = i + 1
f.save("now1.xlsx")  # 保存⽂件