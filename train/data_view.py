import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import xlwt
from sklearn.metrics.pairwise import cosine_similarity
from scipy.spatial.distance import cdist

files = ["shoot/output_shoot_linyl.txt",
         "shoot/output_shoot_xuzy.txt",
         "shoot/output_shoot_xumh.txt",
         "shoot/output_shoot_liujr.txt",
         "shoot/output_shoot_maxd.txt"]
files_test = ["test/output_shoot_linyl.txt",
              "test/output_shoot_xuzy.txt",
              "test/output_shoot_xumh.txt",
              "test/output_shoot_liujr.txt",
              "test/output_shoot_maxd.txt"]
names = ["lyl", "xzy", "xmh", "ljr", "mxd"]
names1 = ["mw1", "mw2", "mw3"]

xlabel = 'time'  # 横坐标标题
ylabel = 'accelerator'  # 纵坐标标题
x_data = [[], [], [], [], []]
y_data = [[], [], [], [], []]
z_data = [[], [], [], [], []]


# 画图显示趋势
def graph(i, title, data_x, data_y, data_z):
    x_alias = list(range(1, len(data_x[0]) + 1))
    sum = 0
    plt.figure()
    for xx in data_x:
        x_alias = list(range(1, len(xx) + 1))
        sum += len(xx) + 1
        draw(title + names[i] + "_x", x_alias, xx)
    plt.show()
    plt.figure()
    for xx in data_y:
        x_alias = list(range(1, len(xx) + 1))
        draw(title + names[i] + "_y", x_alias, xx)
    plt.show()
    plt.figure()
    for xx in data_z:
        x_alias = list(range(1, len(xx) + 1))
        draw(title + names[i] + "_z", x_alias, xx)
    plt.show()
    print("平均长度为", sum / len(data_x))


def draw(title, _x, _y):
    # 画图代码
    # print(data.head()) #可以先看看表的前几行，看有没有载入对
    plt.plot(_x, _y)  # 连线图,若要散点图将此句改为：plt.scatter(x,y) #散点图
    plt.grid(alpha=0.5, linestyle='-.')  # 网格线，更好看
    plt.title(title, fontsize=14)  # 画总标题 fontsize为字体，下同
    plt.xlabel(xlabel, fontsize=14)  # 画横坐标
    plt.ylabel(ylabel, fontsize=14)  # 画纵坐标
    plt.savefig(title + '.jpg', dpi=300)  # 可以存到本地，高清大图。路径默认为当前路径，dpi可理解为清晰度


# 将数据写⼊excel表格
def write2exl(file_path, first_line, datas):
    f = openpyxl.Workbook()
    sheet1 = f.create_sheet(index=0, title="test")  # 创建sheet
    # 将数据写⼊第 i ⾏，第 j 列
    sheet1.append(first_line)
    i = 1
    for data in datas:
        for j in range(len(data)):
            sheet1.cell(i + 1, j + 1, data[j])
        i = i + 1
    f.save(file_path)  # 保存⽂件


# 余弦相似性
def cos_similarity():
    s = cosine_similarity(data_x)
    write2exl("./data_x_cosin.xlsx", s)


# 欧氏距离
def dis():
    distAB = cdist(data_x, data_x, metric='euclidean')
    write2exl("./data_x_dis.xlsx", distAB)


# 求方差
def fc(data_whole):
    fc_whole = []
    for i in range(len(data_whole)):
        sum = 0
        data_ = data_whole[i]
        #print(len(data_whole[0]))
        #print(len(data_whole[0][0]))
        minn = 1000
        for i in range(len(data_whole[0])):
            minn = min(len(data_whole[0][i]),minn)
        data_t = [[data_[i][j]/1000 for i in range(len(data_whole[0]))] for j in range(minn)]
        for arr in data_t:
            arr_var = np.var(arr)
            sum += arr_var
        fc_whole.append(sum/minn)
    return fc_whole


def fc2excel():
    write2exl("./data_fc.xlsx", names, fc_whole)


def get_split(file, index):
    now = []
    data_now = []
    with open(file) as lines:
        for line in lines:
            if len(line) >= 2:
                if line[len(line) - 2] == '-':
                    if len(now) != 0:
                        data_now.append(now)
                    now = []
                    continue
                data = line.split(',')
                now.append(float(data[index]))
    if len(now) != 0:
        data_now.append(now)
    return data_now


# 打开file2，将其附加到who指向的文件得到的data_x,y,z中，绘图，如果ok，（控制台输入），就调用合并函数，将文件合并
def show(file2, who):
    open(files_test[who], 'a')
    data_whole = [[], [], []]
    for index_now in range(3):
        data_whole[index_now] = get_split(files_test[who], index_now)
        data_whole[index_now].extend(get_split(file2, index_now))
    graph(who, "test/", data_whole[0], data_whole[1], data_whole[2])
    s = input("输入y展示现在的方差\n")
    fc_ = fc(data_whole)
    print("x方差", fc_[0])
    print("y方差", fc_[1])
    print("z方差", fc_[2])
    # 获取键盘输入，然后合并
    s = input("请输入y合并\n")
    if s == 'y':
        file_src = open(file2, 'r')
        file_dst = open(files_test[who], 'a+')
        src_tmp = file_src.read()  # 读取源文件内容
        file_dst.write(src_tmp)


def show_origin(file2,who):
    data_whole = [[], [], []]
    for index_now in range(3):
        data_whole[index_now] = get_split(file2, index_now)
    graph(who, "test/", data_whole[0], data_whole[1], data_whole[2])
    fc_ = fc(data_whole)
    print("x方差", fc_[0])
    print("y方差", fc_[1])
    print("z方差", fc_[2])

#
# show_origin("former/output_wing_hyw.txt",0)
# show_origin("former/output_wing_dengyl.txt",1)
# show_origin("former/output_wing_jiangyh.txt",2)
# show_origin("former/output_wing_liucx.txt")
# show_origin("former/output_wing_lsj.txt")

# show_origin("shoot/output_shoot_linyl.txt")
